"""Tracker export endpoint (TED-554).

Generates an .xlsx workbook that mirrors the Team Daily Tracker grid for a
single module: rows = team members, columns = calendar days, each cell the
number of entries that member *created* (bucketed by ``added_at``) on that day,
colored to match the on-screen tracker. Unlike the live grid (which hides the 0
on today/weekends for a clean glance), the export writes the count for *every day
that has occurred* — 0 included — so a downloaded report never has empty cells in
the past (otherwise a 0 on the last day looks like a blank):

    count > 0            -> green   (DCFCE7 / text 15803D), value = count
    past weekday, 0      -> red     (FEE2E2 / text B91C1C), value 0
    today, 0             -> blue    (EEF2FF), value 0
    weekend, 0           -> gray    (F3F4F6), value 0
    future               -> blank   (future weekends shaded gray)

Days are bucketed in the *client's* timezone (passed as ``tz``) so the grid
lines up cell-for-cell with what the browser renders — the frontend tracker
buckets by local ``added_at`` date (see KpiModulePage.tsx TrackerView).
"""
from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

from django.db.models import Count, Q
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from auth_app.models import CustomUser
from roles.permissions import user_is_hod

from .models import (
    GeneralNewEntry,
    GeneralRenewalEntry,
    MarineNewEntry,
    MarineRenewalEntry,
    MedicalClaimEntry,
    MotorClaimEntry,
    MotorFleetNewEntry,
    MotorFleetRenewalEntry,
    MotorNewEntry,
    MotorRenewalEntry,
    SalesKPIEntry,
)

# Module key -> entry model. Keys match the viewset ``module_key`` values.
MODULE_MODELS = {
    'general_new': GeneralNewEntry,
    'general_renewal': GeneralRenewalEntry,
    'motor_new': MotorNewEntry,
    'motor_renewal': MotorRenewalEntry,
    'motor_fleet_new': MotorFleetNewEntry,
    'motor_fleet_renewal': MotorFleetRenewalEntry,
    'motor_claim': MotorClaimEntry,
    'sales_kpi': SalesKPIEntry,
    'marine_new': MarineNewEntry,
    'marine_renewal': MarineRenewalEntry,
    'medical_claim': MedicalClaimEntry,
}

# Human labels for the sheet title / filename.
MODULE_LABELS = {
    'general_new': 'General New',
    'general_renewal': 'General Renewal',
    'motor_new': 'Motor New',
    'motor_renewal': 'Motor Renewal',
    'motor_fleet_new': 'Motor Fleet New',
    'motor_fleet_renewal': 'Motor Fleet Renewal',
    'motor_claim': 'Motor Claim',
    'sales_kpi': 'Deals',
    'marine_new': 'Marine New',
    'marine_renewal': 'Marine Renewal',
    'medical_claim': 'Medical Claim',
}

MAX_RANGE_DAYS = 366

# Tracker palette (see KpiModulePage.tsx TrackerView).
FILL_GREEN = PatternFill('solid', fgColor='DCFCE7')
FILL_RED = PatternFill('solid', fgColor='FEE2E2')
FILL_BLUE = PatternFill('solid', fgColor='EEF2FF')
FILL_GRAY = PatternFill('solid', fgColor='F3F4F6')
FILL_HEADER = PatternFill('solid', fgColor='F9F9F9')
FONT_GREEN = Font(color='15803D', bold=True)
FONT_RED = Font(color='B91C1C', bold=True)
FONT_BOLD = Font(bold=True)
BORDER = Border(*([Side(style='thin', color='E4E4E4')] * 4))
CENTER = Alignment(horizontal='center', vertical='center')

# date.weekday(): Monday=0 .. Sunday=6.
SHORT_DAY = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


def _parse_date(value, field):
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        raise ValidationError({field: 'Expected a date in YYYY-MM-DD format.'})


def _parse_user_ids(raw):
    if not raw:
        return []
    return [int(p) for p in (s.strip() for s in raw.split(',')) if p.isdigit()]


class TrackerExportView(APIView):
    """GET /api/entries/tracker-export/

    Query params:
        module    (required) module key, e.g. ``sales_kpi``
        start     (required) range start, YYYY-MM-DD (local date)
        end       (required) range end, YYYY-MM-DD (local date, inclusive)
        user_ids  (optional) comma-separated member ids; omit for "all members"
        tz        (optional) IANA timezone for day-bucketing, e.g. ``Asia/Dubai``
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        params = request.query_params

        module = params.get('module')
        model = MODULE_MODELS.get(module)
        if model is None:
            raise ValidationError({'module': 'Unknown or unsupported module.'})

        # Module permission — mirror roles.permissions.HasModulePermission.
        if not user.is_staff:
            role = getattr(user, 'role', None)
            if not role or not role.permissions.filter(module=module).exists():
                raise PermissionDenied('You do not have access to this module.')

        start = _parse_date(params.get('start'), 'start')
        end = _parse_date(params.get('end'), 'end')
        if start > end:
            raise ValidationError({'start': 'start must be on or before end.'})
        if (end - start).days + 1 > MAX_RANGE_DAYS:
            raise ValidationError(
                {'end': f'Date range too large (max {MAX_RANGE_DAYS} days).'}
            )

        tz = self._resolve_tz(params.get('tz'))
        today_local = timezone.now().astimezone(tz).date()
        user_ids = _parse_user_ids(params.get('user_ids'))

        # Data visibility — mirror entries.views.BaseEntryViewSet.get_queryset.
        can_see_all = (
            user.is_staff
            or user_is_hod(user)
            or (getattr(user, 'role', None) and user.role.data_visibility == 'all')
        )

        counts = self._entry_counts(model, start, end, tz, user_ids, can_see_all, user)
        members = self._members(module, user_ids, can_see_all, user)

        wb = self._build_workbook(module, members, start, end, today_local, counts)
        buf = BytesIO()
        wb.save(buf)

        filename = f'tracker_{module}_{start.isoformat()}_{end.isoformat()}.xlsx'
        response = HttpResponse(
            buf.getvalue(),
            content_type=(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # ---- helpers ---------------------------------------------------------

    @staticmethod
    def _resolve_tz(tz_name):
        if tz_name:
            try:
                return ZoneInfo(tz_name)
            except Exception:
                pass
        return timezone.get_current_timezone()

    @staticmethod
    def _entry_counts(model, start, end, tz, user_ids, can_see_all, user):
        """Return ``{(user_id, date): count}`` grouped by the entry creator
        (``added_by``) and the local ``added_at`` day."""
        qs = model.objects.all()
        if not can_see_all:
            qs = qs.filter(Q(added_by=user) | Q(on_behalf_of=user))
        # Coarse UTC window (+/- 1 day) so the DB can use the added_at index;
        # the exact local-day filter happens on the annotation below.
        qs = qs.filter(
            added_at__date__gte=start - timedelta(days=1),
            added_at__date__lte=end + timedelta(days=1),
        )
        if user_ids:
            qs = qs.filter(added_by_id__in=user_ids)

        rows = (
            qs.annotate(day=TruncDate('added_at', tzinfo=tz))
            .filter(day__gte=start, day__lte=end)
            .values('added_by_id', 'day')
            .annotate(count=Count('id'))
        )
        return {(r['added_by_id'], r['day']): r['count'] for r in rows}

    @staticmethod
    def _members(module, user_ids, can_see_all, user):
        """Resolve the member rows. The frontend always sends explicit
        ``user_ids`` (the visible members, or the checked subset); the
        membership query is a fallback for direct API use."""
        if not can_see_all:
            return [user]
        qs = CustomUser.objects.filter(is_active=True).select_related('role')
        if user_ids:
            qs = qs.filter(id__in=user_ids)
        else:
            qs = qs.filter(Q(is_staff=True) | Q(role__permissions__module=module))
        return list(qs.distinct().order_by('full_name', 'email', 'id'))

    @staticmethod
    def _build_workbook(module, members, start, end, today_local, counts):
        days = []
        cursor = start
        while cursor <= end:
            days.append(cursor)
            cursor += timedelta(days=1)

        wb = Workbook()
        ws = wb.active
        ws.title = (MODULE_LABELS.get(module, module) or 'Tracker')[:31]

        # Header row: Member | <day columns> | Total
        head = ws.cell(row=1, column=1, value='Member')
        head.font = FONT_BOLD
        head.fill = FILL_HEADER
        head.border = BORDER
        for i, day in enumerate(days):
            c = ws.cell(row=1, column=2 + i, value=f'{SHORT_DAY[day.weekday()]} {day.day}')
            c.font = FONT_BOLD
            c.fill = FILL_HEADER
            c.alignment = CENTER
            c.border = BORDER
        total_col = 2 + len(days)
        tc = ws.cell(row=1, column=total_col, value='Total')
        tc.font = FONT_BOLD
        tc.fill = FILL_HEADER
        tc.alignment = CENTER
        tc.border = BORDER

        # Body rows.
        for r, member in enumerate(members, start=2):
            name_cell = ws.cell(
                row=r, column=1, value=member.get_full_name() or member.email
            )
            name_cell.border = BORDER
            row_total = 0
            for i, day in enumerate(days):
                count = counts.get((member.id, day), 0)
                cell = ws.cell(row=r, column=2 + i)
                cell.alignment = CENTER
                cell.border = BORDER
                is_weekend = day.weekday() >= 5
                if day > today_local:
                    # Future days have no value; keep weekends shaded for consistency.
                    if is_weekend:
                        cell.fill = FILL_GRAY
                    continue
                # The day has occurred (past or today) -> always write the count,
                # 0 included. The live grid hides 0 on today/weekends for a clean
                # glance, but a downloaded report should show every occurred day
                # (otherwise a 0 on the last day looks like an empty cell).
                cell.value = count
                if count > 0:
                    cell.fill = FILL_GREEN
                    cell.font = FONT_GREEN
                    row_total += count
                elif day == today_local:
                    cell.fill = FILL_BLUE
                elif is_weekend:
                    cell.fill = FILL_GRAY
                else:
                    cell.fill = FILL_RED
                    cell.font = FONT_RED
            total_cell = ws.cell(row=r, column=total_col, value=row_total)
            total_cell.alignment = CENTER
            total_cell.border = BORDER
            total_cell.font = FONT_BOLD

        # Column widths + freeze the member column and header row.
        ws.column_dimensions['A'].width = 26
        for i in range(len(days)):
            ws.column_dimensions[get_column_letter(2 + i)].width = 5
        ws.column_dimensions[get_column_letter(total_col)].width = 8
        ws.freeze_panes = 'B2'
        return wb
